"""
 You need to login to the Bloomberg Terminal for the script to work!
 Run it using the arrow on the top right.
 Enter the stock ticker while specifying the country in the end
 For example AAPL US or 000660 KS (the script will automatically add [Equity])

 On this model there is an issue, no peers are fetched and it subprocess of data is dependant on the terminal so, outside the terminal as they are hardcoded
 the formulas wont work, to fix this, enter the peers name (e.g. AMZN US Equity, etc...) copy the entire data on the peers and special paste it values only
 on the same space it was, shouldnt be any formulas left
 ... AND as excel 2016 doesnt work well with xlookup, you need to download it and open in teams/web/ a pc with excel >2016...

 On the excel if the cell has N/A (or some iteration of this...) an error has occured with the fetching of data and must be checked manually, this is so that
 if for example it had returned 0 instead, it would cause doubt as in some cases the values for certain names are 0

 #--------#-------# TO DO LIST
 To fix... excel peers ratios are not linked, gonna kms before i do that tho
 Peers thing on.. my pc doesnt do that idk why 
 Holders and board, must also be same issue, double check on 23

 https://www.youtube.com/watch?v=qxgSbP3H3YY

 PARA DE APAGAR OS MEUS FICHEIROS
"""
import blpapi
import openpyxl
import shutil
import os
import time
from datetime import datetime

def setup_bloomberg_session(ticker_symbol):
    options = blpapi.SessionOptions()
    options.setServerHost("localhost")
    options.setServerPort(8194)
    session = blpapi.Session(options)

    print(f"\n✨ Attempting to connect to Bloomberg for {ticker_symbol}...")
    if not session.start():
        print("⚠️ Whoops! Failed to start the Bloomberg session. Please make sure the Bloomberg Terminal is running and the API is enabled.")
        return None
    if not session.openService("//blp/refdata"):
        print("⚠️ Uh oh! Failed to open the Bloomberg reference data service. Can't get the data without this.")
        session.stop()
        return None
    print(f"✅ Great! Bloomberg session for {ticker_symbol} started successfully. Let's get some data!")
    return session

def fetch_bloomberg_data(session, ticker, fields, field_to_name_map, start_year=2014, end_year=2024, timeout=30):
    if not fields:
        print("ℹ️ Just a heads up: No specific data fields were requested for this round.")
        return {}

    if len(fields) > 25:
        print(f"🔍 Lots of data requested! ({len(fields)} fields). Don't worry, I'll handle it, but if this were a single request, it might be a bit much for Bloomberg.")

    ref_data_service = session.getService("//blp/refdata")
    request = ref_data_service.createRequest("HistoricalDataRequest")

    security_for_request = f"{ticker} Equity"
    request.getElement("securities").appendValue(security_for_request)

    parts = ticker.strip().split()
    country_code = ""
    if len(parts) > 1 and len(parts[-1]) == 2 and parts[-1].isalpha():
        country_code = parts[-1].upper()

    if country_code and country_code != "US":
        print(f"🌍 Looks like a non-US stock ({ticker}, Country: {country_code}). I'll ask Bloomberg for data in USD to keep things consistent.")
        
        request.set("currency", "USD")

        # USD currency override
        #overrides = request.getElement("overrides")
        #override = overrides.appendElement()
        #override.setElement("fieldId", "EQY_FUND_CRNCY")
        #override.setElement("value", "USD")
        
    elif country_code == "US":
        print(f"🇺🇸 This stock ({ticker}) is US-based. Data should come in USD by default.")
    else:
        print(f"🤔 Can't quite tell the country for {ticker} from the ticker. I'll request data in its local currency.")

    for field in fields:
        request.getElement("fields").appendValue(field)

    request.set("periodicitySelection", "YEARLY")
    request.set("startDate", f"{start_year}0101")
    request.set("endDate", f"{end_year}1231")

    print(f"📡 Sending a request to Bloomberg for {security_for_request}. Asking for: {fields}")
    session.sendRequest(request)

    data = {field: {} for field in fields}
    invalid_fields = []
    start_time = time.time()

    while True:
        if time.time() - start_time >= timeout:
            print(f"⏳ Timeout! It's been {timeout} seconds, and I'm still waiting for all the data for {security_for_request}. I'll mark the missing pieces as 'N/A (Timeout)'.")
            for field_id_timeout in fields:
                if field_id_timeout not in invalid_fields:
                    for year_timeout_val in range(start_year, end_year + 1):
                        if year_timeout_val not in data.get(field_id_timeout, {}):
                            data.setdefault(field_id_timeout, {})[year_timeout_val] = "N/A (Timeout)"
            break

        event = session.nextEvent(500)

        if event.eventType() == blpapi.Event.TIMEOUT:
            print(f"🕰️ Bloomberg is taking a moment to respond for {security_for_request}. Still waiting...")
            continue

        for msg in event:

            if msg.hasElement("responseError"):
                error = msg.getElement("responseError")
                error_message = error.getElement("message").getValue()
                print(f"❌ Error from Bloomberg for {security_for_request}: {error_message}. This might affect the data for the current request.")
                for f_id_err in fields:
                    if f_id_err not in invalid_fields: invalid_fields.append(f_id_err)
                continue

            if not msg.hasElement("securityData"):
                print(f"🤔 Hmm, a message came from Bloomberg for {security_for_request}, but it's missing the main 'securityData' part. That's odd.")
                continue

            security_data = msg.getElement("securityData")

            if security_data.hasElement("securityError"):
                sec_error = security_data.getElement("securityError")
                error_msg_sec = sec_error.getElement("message").getValueAsString()
                print(f"❌ Security Error from Bloomberg for '{security_data.getElementValue('security')}': {error_msg_sec}. This means I can't get data for this stock with the current fields.")
                for f_id_sec_err in fields:
                    if f_id_sec_err not in invalid_fields: invalid_fields.append(f_id_sec_err)
                continue

            if security_data.hasElement("fieldExceptions"):
                field_exceptions = security_data.getElement("fieldExceptions")
                for j in range(field_exceptions.numValues()):
                    field_error = field_exceptions.getValue(j)
                    invalid_field_id = field_error.getElement("fieldId").getValueAsString()
                    error_info = field_error.getElement("errorInfo")
                    error_message_field = error_info.getElement("message").getValueAsString()
                    field_name_display = field_to_name_map.get(invalid_field_id, "Unknown Field (Not in my list)")
                    print(f"⚠️ Problem with a specific field: '{invalid_field_id}' (which I know as '{field_name_display}') for {security_for_request}. Bloomberg says: {error_message_field}. I'll mark it as N/A.")
                    if invalid_field_id not in invalid_fields:
                        invalid_fields.append(invalid_field_id)

            if not security_data.hasElement("fieldData"):
                print(f"😕 No actual data ('fieldData') found in the message from Bloomberg for {security_for_request}. This can happen if all requested fields were invalid or there's no data for the chosen period.")
                for f_id_no_data in fields:
                    if f_id_no_data not in invalid_fields:
                        for year_val_nd in range(start_year, end_year + 1):
                            data.setdefault(f_id_no_data, {})[year_val_nd] = "N/A (No data found)"
                continue

            field_data_array = security_data.getElement("fieldData")
            
            for k in range(field_data_array.numValues()):
                datum = field_data_array.getValue(k)
                if not datum.hasElement("date"):
                    print("🤔 One of the data entries from Bloomberg is missing a 'date'. Skipping this particular entry.")
                    continue
                date_obj = datum.getElement("date").getValueAsDatetime()
                year = date_obj.year

                for field_id in fields:
                    if field_id in invalid_fields:
                        data.setdefault(field_id, {})[year] = "N/A (Invalid Field)"
                        continue

                    if datum.hasElement(field_id):
                        try:
                            value = datum.getElement(field_id).getValueAsFloat()
                            data.setdefault(field_id, {})[year] = value
                        except blpapi.exception.ElementErrorException:
                            try:
                                value_str = datum.getElement(field_id).getValueAsString()
                                data.setdefault(field_id, {})[year] = value_str
                                print(f"📝 Note: The data for field '{field_id}' for year {year} for {security_for_request} isn't a number. I've saved it as text: '{value_str}'")
                            except Exception as e_str:
                                print(f"⚠️ Couldn't get the value for field '{field_id}' for year {year} for {security_for_request}, even as text. Error: {e_str}. Marking as N/A.")
                                data.setdefault(field_id, {})[year] = "N/A (Error reading)"
                        except Exception as e_gen:
                            print(f"⚠️ Error trying to read field '{field_id}' for year {year} for {security_for_request}: {e_gen}. Marking as N/A.")
                            data.setdefault(field_id, {})[year] = "N/A (Conversion Error)"
                    else:
                        if year not in data.get(field_id, {}): 
                            data.setdefault(field_id, {})[year] = None 

        if event.eventType() == blpapi.Event.RESPONSE:
            print(f"📬 Received the complete response from Bloomberg for this batch ({security_for_request}).")
            for field_id_fill in fields:
                if field_id_fill not in invalid_fields: 
                    for year_fill_val in range(start_year, end_year + 1):
                        if year_fill_val not in data.get(field_id_fill, {}): 
                            data.setdefault(field_id_fill, {})[year_fill_val] = "N/A (Missing)"
            break

        elif event.eventType() in [blpapi.Event.SESSION_STATUS, blpapi.Event.SERVICE_STATUS]:
            for msg in event:
                if msg.messageType() == blpapi.Name("SessionTerminated"):
                    print("❗CRITICAL: The Bloomberg session was terminated unexpectedly! Further data requests might fail.")
                    return None 
                print(f"ℹ️ Just an update from Bloomberg (Session/Service Status): {msg.toString().strip()}")
        else:
            print(f"🤖 Received an unusual event type from Bloomberg: {event.eventType()}. Message: {msg.toString().strip()}")

    if not any(data.get(field) for field in data):
        print(f"💨 It seems no data was successfully retrieved for any requested field for {ticker} in this batch.")

    if invalid_fields:
        print(f"🚫 For {security_for_request}, these Bloomberg fields were skipped or marked N/A because of errors or invalidity: {invalid_fields}")

    return data

def calculate_derived_metrics(fetched_data, start_year, end_year):
    derived_data = {}
    
    other_operating_components = [
        "IS_OTHER_OPER_INC",  
        "IS_OTHER_OPERATING_EXPENSES",
        "IS_FOREIGN_EXCH_LOSS",
        "INCOME_LOSS_FROM_AFFILIATES",
        "IS_ABNORMAL_ITEM"
        #"OTHER_NONOP_INCOME"
    ]

    derived_data["Total_Other_Operating"] = {}
    
    for year in range(start_year, end_year + 1):
        total_other = 0
        has_value = False
        
        for field in other_operating_components:
            if field in fetched_data and year in fetched_data[field]:
                value = fetched_data[field].get(year)
                if isinstance(value, (int, float)):
                    total_other += value
                    has_value = True
        
        
        if has_value:
            derived_data["Total_Other_Operating"][year] = total_other
        else:
            derived_data["Total_Other_Operating"][year] = "N/A (Missing Components)"

    return derived_data

field_map = {
    "Revenue (Sales)": {"source": "BDH", "field": "SALES_REV_TURN", "statement": "IS"},
    "COGS (Cost of Goods Sold)": {"source": "BDH", "field": "IS_COG_AND_SERVICES_SOLD", "statement": "IS"},
    "Gross Profit": {"source": "BDH", "field": "GROSS_PROFIT", "statement": "IS"},
    "SG&A (Selling, General & Administrative)": {"source": "BDH", "field": "IS_SGA_EXPENSE", "statement": "IS"},
    "R&D (Research & Development)": {"source": "BDH", "field": "IS_OPERATING_EXPENSES_RD", "statement": "IS"},
    "Other Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_OPER_INC", "statement": "IS"},
    "EBITDA": {"source": "BDH", "field": "EBITDA", "statement": "IS"},
    "D&A (Depreciation & Amortization)": {"source": "BDH", "field": "CF_DEPR_AMORT", "statement": "CF"},
    "Depreciation Expense": {"source": "BDH", "field": "ARDR_DEPRECIATION_EXP", "statement": "IS"},
    "Amortization Expense": {"source": "BDH", "field": "ARDR_AMORT_EXP", "statement": "IS"},
    "Operating Income (EBIT)": {"source": "BDH", "field": "IS_OPER_INC", "statement": "IS"},
    "Net Interest Expense (Income)": {"source": "BDH", "field": "IS_NET_INTEREST_EXPENSE", "statement": "IS"},
    "Interest Expense": {"source": "BDH", "field": "IS_INT_EXPENSE", "statement": "IS"},
    "Interest Income": {"source": "BDH", "field": "IS_INT_INC", "statement": "IS"},
    "FX (Gain) Loss": {"source": "BDH", "field": "IS_FOREIGN_EXCH_LOSS", "statement": "IS"},
    "Other Non-Operating (Income) Expenses": {"source": "BDH", "field": "IS_OTHER_NON_OPERATING_INC_LOSS", "statement": "IS"},
    "Pre-Tax Income (EBT)": {"source": "BDH", "field": "PRETAX_INC", "statement": "IS"},
    "Tax Expense (Benefits)": {"source": "BDH", "field": "IS_INC_TAX_EXP", "statement": "IS"},
    "Net Income": {"source": "BDH", "field": "NET_INCOME", "statement": "IS"},
    "EPS Basic": {"source": "BDH", "field": "BASIC_EPS", "statement": "IS"},
    "EPS Diluted": {"source": "BDH", "field": "DILUTED_EPS", "statement": "IS"},
    "Basic Weighted Average Shares": {"source": "BDH", "field": "IS_AVG_NUM_SH_FOR_EPS", "statement": "IS"},
    "Diluted Weighted Average Shares": {"source": "BDH", "field": "IS_SH_FOR_DILUTED_EPS", "statement": "IS"},
    "Total Other Operating Components": {"source": "derived", "field": "Total_Other_Operating", "statement": "IS"},

    "Cash & Cash Equivalents": {"source": "BDH", "field": "BS_CASH_NEAR_CASH_ITEM", "statement": "BS"},
    "Short-Term Investments": {"source": "BDH", "field": "BS_MKT_SEC_OTHER_ST_INVEST", "statement": "BS"},
    "Accounts Receivable": {"source": "BDH", "field": "BS_ACCT_NOTE_RCV", "statement": "BS"},
    "Inventory": {"source": "BDH", "field": "BS_INVENTORIES", "statement": "BS"},
    "Current Assets": {"source": "BDH", "field": "BS_CUR_ASSET_REPORT", "statement": "BS"},
    "Gross PP&E (Property, Plant and Equipment)": {"source": "BDH", "field": "BS_GROSS_FIX_ASSET", "statement": "BS"},
    "Accumulated Depreciation": {"source": "BDH", "field": "BS_ACCUM_DEPR", "statement": "BS"},
    "Intangibles": {"source": "BDH", "field": "BS_DISCLOSED_INTANGIBLES", "statement": "BS"},
    "Goodwill": {"source": "BDH", "field": "BS_GOODWILL", "statement": "BS"},
    "Non-Current Assets": {"source": "BDH", "field": "BS_TOT_NON_CUR_ASSET", "statement": "BS"},
    "Accounts Payable": {"source": "BDH", "field": "BS_ACCT_PAYABLE", "statement": "BS"},
    "Short-Term Borrowings": {"source": "BDH", "field": "SHORT_TERM_DEBT_DETAILED", "statement": "BS"},
    "Current Portion of Lease Liabilities": {"source": "BDH", "field": "ST_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    "Current Liabilities": {"source": "BDH", "field": "BS_CUR_LIAB", "statement": "BS"},
    "Long-Term Borrowings": {"source": "BDH", "field": "LONG_TERM_BORROWINGS_DETAILED", "statement": "BS"},
    "Long-Term Operating Lease Liabilities": {"source": "BDH", "field": "LT_CAPITALIZED_LEASE_LIABILITIES", "statement": "BS"},
    "Non-Current Liabilities": {"source": "BDH", "field": "NON_CUR_LIAB", "statement": "BS"},
    "Non-Controlling Interest": {"source": "BDH", "field": "MINORITY_NONCONTROLLING_INTEREST", "statement": "BS"},
    "Right-of-Use Assets": {"source": "BDH", "field": "TOT_OPER_LEA_RT_OF_USE_ASSETS", "statement": "BS"},

    "(Increase) Decrease in Accounts Receivable": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Inventories": {"source": "BDH", "field": "CF_CHANGE_IN_INVENTORIES", "statement": "CF", "section": "Operating"},
    "(Increase) Decrease in Pre-paid expeses and Other CA": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "Increase (Decrease) in Accounts Payable": {"source": "BDH", "field": "CF_CHANGE_IN_ACCOUNTS_PAYABLE", "statement": "CF", "section": "Operating"},
    "Increase (Decrease) in Accrued Revenues and Other CL": {"source": "BDH", "field": "CF_ACCT_RCV_UNBILLED_REV", "statement": "CF", "section": "Operating"},
    "Stock Based Compensation": {"source": "BDH", "field": "CF_STOCK_BASED_COMPENSATION", "statement": "CF", "section": "Operating"},
    "Operating Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_OPER", "statement": "CF", "section": "Operating"},
    "Acquisition of Fixed & Intangibles": {"source": "BDH", "field": "ACQUIS_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"},
    "Disposal of Fixed & Intangibles": {"source": "BDH", "field": "DISPOSAL_OF_FIXED_INTANG", "statement": "CF", "section": "Investing"},
    "Acquisitions": {"source": "BDH", "field": "CF_CASH_FOR_ACQUIS_SUBSIDIARIES", "statement": "CF", "section": "Investing"},
    "Divestitures": {"source": "BDH", "field": "CF_CASH_FOR_DIVESTITURES", "statement": "CF", "section": "Investing"},
    "Increase in LT Investment": {"source": "BDH", "field": "CF_INCR_INVEST", "statement": "CF", "section": "Investing"},
    "Decrease in LT Investment": {"source": "BDH", "field": "CF_DECR_INVEST", "statement": "CF", "section": "Investing"},
    "Investing Cash Flow": {"source": "BDH", "field": "CF_CASH_FROM_INV_ACT", "statement": "CF", "section": "Investing"},
    "Lease Payments": {"source": "BDH", "field": "ARDR_REPAYMENT_FINANCE_LEASES", "statement": "CF", "section": "Financing"},
    "Debt Borrowing": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PROCEEDS", "statement": "CF", "section": "Financing"},
    "Debt Repayment": {"source": "BDH", "field": "CF_LT_DEBT_CAP_LEAS_PAYMENT", "statement": "CF", "section": "Financing"},
    "Dividends": {"source": "BDH", "field": "CF_DVD_PAID", "statement": "CF", "section": "Financing"},
    "Increase (Repurchase) of Shares": {"source": "BDH", "field": "PROC_FR_REPURCH_EQTY_DETAILED", "statement": "CF", "section": "Financing"},
    "Financing Cash Flow": {"source": "BDH", "field": "CFF_ACTIVITIES_DETAILED", "statement": "CF", "section": "Financing"},
    "Effect of Foreign Exchange": {"source": "BDH", "field": "CF_EFFECT_FOREIGN_EXCHANGES", "statement": "CF", "section": "All"},
    #"Net Changes in Cash": {"source": "BDH", "field": "CF_NET_CHNG_CASH", "statement": "CF", "section": "All"},

    "Market Capitalization": {"source": "BDH", "field": "CUR_MKT_CAP", "statement": "BS"},
    "Total Debt": {"source": "BDH", "field": "SHORT_AND_LONG_TERM_DEBT", "statement": "BS"},
    "Preferred Stock": {"source": "BDH", "field": "PFD_EQTY_HYBRID_CAPITAL", "statement": "BS"},
    "Enterprise Value": {"source": "BDH", "field": "ENTERPRISE_VALUE", "statement": "BS"},
}

field_cell_map = {
    "Revenue (Sales)": "G6",
    "COGS (Cost of Goods Sold)": "G7",
    "Gross Profit": "G8",
    "SG&A (Selling, General & Administrative)": "G9",
    "R&D (Research & Development)": "G10",
    "Total Other Operating Components": "G11",
    "EBITDA": "G12",
    "D&A (Depreciation & Amortization)": "G13",
    "Depreciation Expense": "G14",
    "Amortization Expense": "G15",
    "Operating Income (EBIT)": "G16",
    "Net Interest Expense (Income)": "G17",
    "Interest Expense": "G18",
    "Interest Income": "G19",
    "FX (Gain) Loss": "G20",
    "Other Non-Operating (Income) Expenses": "G21",
    "Pre-Tax Income (EBT)": "G22",
    "Tax Expense (Benefits)": "G23",
    "Net Income": "G24",
    "EPS Basic": "G25",
    "EPS Diluted": "G26",
    "Basic Weighted Average Shares": "G27",
    "Diluted Weighted Average Shares": "G28",

    "Cash & Cash Equivalents": "G33",
    "Short-Term Investments": "G34",
    "Accounts Receivable": "G35",
    "Inventory": "G36",
    "Current Assets": "G38",
    "Gross PP&E (Property, Plant and Equipment)": "G40",
    "Accumulated Depreciation": "G41",
    "Right-of-Use Assets": "G42",
    "Intangibles": "G43",
    "Goodwill": "G44",
    "Non-Current Assets": "G47",
    "Accounts Payable": "G49",
    "Short-Term Borrowings": "G51",
    "Current Portion of Lease Liabilities": "G52",
    "Current Liabilities": "G54", 
    "Long-Term Borrowings": "G56",
    "Long-Term Operating Lease Liabilities": "G57",
    "Non-Current Liabilities": "G59", 
    "Non-Controlling Interest": "G62",

    "(Increase) Decrease in Accounts Receivable": "G69",
    "(Increase) Decrease in Inventories": "G70",
    "(Increase) Decrease in Pre-paid expeses and Other CA": "G71",
    "Increase (Decrease) in Accounts Payable": "G72",
    "Increase (Decrease) in Accrued Revenues and Other CL": "G73",
    "Stock Based Compensation": "G74",
    "Operating Cash Flow": "G76",
    "Acquisition of Fixed & Intangibles": "G78",
    "Disposal of Fixed & Intangibles": "G79", 
    "Acquisitions": "G81",
    "Divestitures": "G82",
    "Increase in LT Investment": "G83",
    "Decrease in LT Investment": "G84",
    "Investing Cash Flow": "G86",
    "Lease Payments": "G89",
    "Debt Borrowing": "G87", 
    "Debt Repayment": "G88", 
    "Dividends": "G90",
    "Increase (Repurchase) of Shares": "G91",
    "Financing Cash Flow": "G93",
    "Effect of Foreign Exchange": "G94",
    #"Net Changes in Cash": "G95",

    "Market Capitalization": "G99",
    "Total Debt": "G101",
    "Preferred Stock": "G102",
    "Enterprise Value": "G104",
}


def filter_field_map_for_task(task_name, current_field_map):
    statement_code = task_name

    allowed_statements = ["IS", "BS", "CF"]
    if statement_code not in allowed_statements:
        raise ValueError(f"Error: Invalid statement code '{statement_code}'. Must be one of {allowed_statements}.")

    task_specific_configs = {}
    for name, config in current_field_map.items():
        if config["statement"] == statement_code:
            task_specific_configs[name] = config

    required_bdh_for_derived_metrics = set()

    for bdh_field_code_needed in required_bdh_for_derived_metrics:
        found_in_task = False
        for _, existing_config in task_specific_configs.items():
            if existing_config.get("field") == bdh_field_code_needed and existing_config.get("source") == "BDH":
                found_in_task = True
                break
        if not found_in_task:
            pass
    return task_specific_configs


def batch_fields(fields_to_fetch, batch_size=25):
    unique_fields = sorted(list(set(fields_to_fetch))) 
    return [unique_fields[i:i + batch_size] for i in range(0, len(unique_fields), batch_size)]

def get_column_letter_from_index(col_index):
    return openpyxl.utils.get_column_letter(col_index)

def get_target_cells_for_years(base_cell_ref, num_years):
    try:
        col_str = "".join(filter(str.isalpha, base_cell_ref))
        row_num = int("".join(filter(str.isdigit, base_cell_ref)))
        start_col_idx = openpyxl.utils.column_index_from_string(col_str)

        target_cells = []
        for i in range(num_years):
            target_col_letter = get_column_letter_from_index(start_col_idx + i)
            target_cells.append(f"{target_col_letter}{row_num}")
        return target_cells
    except ValueError as e:
        print(f"❌ Error: The cell reference '{base_cell_ref}' looks wrong. Could not figure out the column and row: {e}")
        raise 

def populate_valuation_model(template_path, output_path, ticker_symbol, current_field_map, current_field_cell_map):
    if not os.path.exists(template_path):
        print(f"❌ Oh no! The Excel template file '{template_path}' wasn't found. Please make sure it's in the same directory as the script or you've provided the full path.")
        raise FileNotFoundError(f"Template file {template_path} not found.")

    try:
        shutil.copy(template_path, output_path)
        print(f"📄 Copied the template '{template_path}' to your new output file '{output_path}'.")
    except Exception as e_copy:
        print(f"❌ Failed to copy the template to '{output_path}'. Error: {e_copy}")
        raise

    try:
        wb = openpyxl.load_workbook(output_path)
    except Exception as e_load:
        print(f"❌ Trouble opening the new Excel file '{output_path}'. Error: {e_load}")
        raise

    try:
        if "DCF" in wb.sheetnames:
            dcf_sheet = wb["DCF"]
            dcf_sheet["R4"] = ticker_symbol 
            print(f"📝 Ticker '{ticker_symbol}' written to cell R4 of 'DCF' sheet.")
        else:
            print(f"⚠️ Sheet 'DCF' not found in the workbook '{output_path}'. Ticker '{ticker_symbol}' NOT written to cell R4.")
    except Exception as e:
        print("An erroe occured with the message: ", e)
            
    if "Inputs" not in wb.sheetnames:
        print("❌ The Excel template is missing the 'Inputs' sheet. I need that sheet to put the data in! Please check the template.")
        raise ValueError("'Inputs' sheet not found in the template file.")
    ws = wb["Inputs"]

    data_years = list(range(2014, 2024 + 1))
    num_data_years = len(data_years)

    all_fetched_bdh_data = {} 

    global_bberg_code_to_excel_name_map = {
        config["field"]: name
        for name, config in current_field_map.items()
        if config.get("source") == "BDH" and "field" in config
    }

    all_bdh_fields_to_fetch_codes = set()
    for excel_name, config in current_field_map.items():
        if config.get("source") == "BDH" and "field" in config:
            all_bdh_fields_to_fetch_codes.add(config["field"])
        elif config.get("source") == "derived": 
            pass


    if not all_bdh_fields_to_fetch_codes:
        print("🤔 It seems no Bloomberg data fields (BDH fields) are listed in the configuration. I can't fetch anything without them. Please check the 'field_map'.")
        wb.save(output_path) 
        return

    print(f"\n🚀 Phase 1: Starting data hunt for ticker: {ticker_symbol}")
    print(f"📊 I need to find {len(all_bdh_fields_to_fetch_codes)} unique pieces of data from Bloomberg.")

    field_batches = batch_fields(list(all_bdh_fields_to_fetch_codes), batch_size=25)
    print(f"📦 I've split this into {len(field_batches)} smaller batches to ask Bloomberg.")

    session = None 
    try:
        session = setup_bloomberg_session(ticker_symbol)
        if not session:
            print(f"❌ Major setback: Failed to start the Bloomberg session for {ticker_symbol}. I can't fetch any data. Please check your Bloomberg Terminal connection.")
            
            raise ConnectionError("Failed to establish Bloomberg session.")

        for batch_idx, current_batch_bberg_codes in enumerate(field_batches):
            print(f"    🔎 Batch {batch_idx + 1} of {len(field_batches)}: Asking for {len(current_batch_bberg_codes)} specific items.")

            batch_data_fetched = fetch_bloomberg_data(
                session,
                ticker_symbol,
                current_batch_bberg_codes,
                global_bberg_code_to_excel_name_map,
                start_year=data_years[0],
                end_year=data_years[-1]
            )

            if batch_data_fetched is None: 
                print(f"    ❗ Critical Error: Something went wrong with the Bloomberg connection during batch {batch_idx + 1}. Stopping further data fetching.")
                raise ConnectionAbortedError("Bloomberg session terminated or critical fetch error during a batch.")


            elif batch_data_fetched: 
                for field_code, yearly_data in batch_data_fetched.items():
                    if field_code not in all_fetched_bdh_data:
                        all_fetched_bdh_data[field_code] = {}
                    for year, value in yearly_data.items():
                        if value is not None:
                            all_fetched_bdh_data[field_code][year] = value
                        elif year not in all_fetched_bdh_data[field_code]:
                            all_fetched_bdh_data[field_code][year] = value 

                print(f"    👍 Success! Got data for batch {batch_idx + 1}. Processed {len(batch_data_fetched)} field types from this batch.")
            else:
                print(f"    ℹ️ Batch {batch_idx + 1} didn't return any data. This could be because all fields in it were invalid or no data was available.")

    except ConnectionError as e_conn_err: 
        print(f"❌ Connection Error: {e_conn_err}")
    except ConnectionAbortedError as e_conn_abort: 
        print(f"❌ Connection Aborted: {e_conn_abort}")
    except Exception as e_fetch:
        print(f"❌ An unexpected error occurred while trying to get data from Bloomberg: {e_fetch}")

    finally:
        if session:
            try:
                session.stop()
                print("🔌 Bloomberg session stopped. All done with data fetching (or tried our best!).")
            except Exception as e_stop:
                print(f"⚠️ Minor issue while trying to stop the Bloomberg session: {e_stop}")

    print(f"\n🏁 Phase 1 Complete: Finished all attempts to fetch data from Bloomberg.")

    print(f"\n🧮 Phase 2: Calculating any extra metrics (like DSO)...")
    all_derived_data = calculate_derived_metrics(all_fetched_bdh_data, start_year=data_years[0], end_year=data_years[-1])
    print("✅ Derived metrics calculated (if any were defined).")

    print(f"\n✍️ Phase 3: Writing all the gathered data into your Excel sheet: '{ws.title}'...")

    for excel_name, config in current_field_map.items():
        if excel_name.startswith("__dep_"): 
            continue

        base_cell_ref = current_field_cell_map.get(excel_name)
        if not base_cell_ref:
            print(f"🤔 Couldn't find where to put '{excel_name}' in the Excel sheet (no cell mapping). Skipping this item.")
            continue

        try:
            target_cells_for_item = get_target_cells_for_years(base_cell_ref, num_data_years)
        except Exception as e_cell_calc:
            print(f"❌ Error figuring out the cells for '{excel_name}' (starting from '{base_cell_ref}'): {e_cell_calc}. Skipping this one.")
            continue

        data_source_for_item = {}
        source_type = config.get("source", "unknown").upper()

        if source_type == "BDH":
            bberg_field_code = config.get("field")
            if not bberg_field_code:
                print(f"🤔 The item '{excel_name}' is marked as Bloomberg data (BDH), but has no Bloomberg field code. Skipping.")
                continue
            data_source_for_item = all_fetched_bdh_data.get(bberg_field_code, {})
            if not data_source_for_item:
                print(f"💨 No data was fetched for '{excel_name}' (Bloomberg code: {bberg_field_code}). It will be marked N/A.")


        elif source_type == "DERIVED":
            derived_field_key = config.get("field") 
            if not derived_field_key:
                print(f"🤔 The item '{excel_name}' is marked as 'derived', but I don't know which calculation it refers to. Skipping.")
                continue
            data_source_for_item = all_derived_data.get(derived_field_key, {})
            if not data_source_for_item:
                print(f"💨 No data was calculated for the derived metric '{excel_name}'. It will be marked N/A (likely due to missing inputs).")

        else:
            print(f"❓ Item '{excel_name}' has an unknown data source type: '{config.get('source', 'Not Specified')}'. Skipping.")
            continue

        for i, year in enumerate(data_years):
            cell_ref = target_cells_for_item[i]
            raw_value = data_source_for_item.get(year)

            display_value = raw_value 
            #skibidi toilet
            
            if raw_value is None: 
                display_value = "0" 

            try:
                if isinstance(raw_value, (int, float)):
                    ws[cell_ref] = raw_value
                    ws[cell_ref].number_format = "#,##0.000" 
                    if "EPS" in excel_name or "DSO" in excel_name or "Rate" in excel_name: 
                        ws[cell_ref].number_format = "0.00"
                elif isinstance(raw_value, str): 
                    ws[cell_ref] = raw_value
                else: 
                    ws[cell_ref] = "0" if raw_value is None else str(raw_value) 
            except Exception as e_write_cell:
                print(f"⚠️ Problem writing to cell {cell_ref} for '{excel_name}': {e_write_cell}")
                ws[cell_ref] = "Error writing"


    try:
        wb.save(output_path)
        print(f"\n🎉 All Done! Your valuation model has been populated and saved to: '{output_path}'")
    except Exception as e_save:
        print(f"❌ Critical Error: Failed to save the final Excel workbook to '{output_path}'. Error: {e_save}")
        print("    Possible reasons: The file might be open in Excel, or there might be a permissions issue with the folder.")


if __name__ == "__main__":
    print("-" * 70)
    print(" ✨ Bloomberg Data to Excel Valuation Model Populator ✨ ")
    print("-" * 70)

    excel_template_path = "LIS_Valuation_Empty.xlsx" 

    try:
        desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")
        output_folder_name = "Bloomberg_Valuation_Reports"
        output_directory = os.path.join(desktop_path, output_folder_name)

        if not os.path.exists(output_directory):
            os.makedirs(output_directory)
            print(f"📂 Created a new folder for your reports on your Desktop: '{output_directory}'")
        else:
            print(f"📂 Using the existing folder for reports on your Desktop: '{output_directory}'")
    except Exception as e_path:
        print(f"⚠️ Couldn't create or access the Desktop folder for reports (Error: {e_path}).")
        print("    I'll save the report in the same directory as this script instead.")
        output_directory = "." 

    ticker_input = ""
    while not ticker_input:
        raw_input_str = input("➡️ Enter the Ticker Symbol (e.g., AAPL US or 000660 KS): ").strip()
        if raw_input_str and any(char.isalnum() for char in raw_input_str):
            ticker_input = raw_input_str.upper() 
        else:
            print("❗ Please enter a valid ticker symbol. It can't be empty and should have some letters or numbers.")

    safe_ticker_filename = ticker_input.replace(" ", "_").replace("/", "_")
    output_file_name = f"{safe_ticker_filename}_Valuation_Model_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx" # Corrected strftime
    final_output_path = os.path.join(output_directory, output_file_name)

    print(f"\n📝 Using Template: '{excel_template_path}'")
    print(f"💾 Final Report Will Be Saved As: '{final_output_path}'")
    print(f"🎯 Ticker for Bloomberg: '{ticker_input}' (Script automatically adds 'Equity' during request)") # Clarified ticker usage

    try:
        print("\n⏳ Starting the data population process... This might take a few moments.\n")
        populate_valuation_model(
            template_path=excel_template_path,
            output_path=final_output_path,
            ticker_symbol=ticker_input,
            current_field_map=field_map, 
            current_field_cell_map=field_cell_map
        )
    
        print("\n✅ Process completed successfully!")
    except FileNotFoundError as e_fnf:
        print(f"❌ CRITICAL ERROR: The Excel template file was not found. {e_fnf}")
        print("    Please make sure the template file ('LIS_Valuation_Empty.xlsx') is in the same directory as this script.")
    except ConnectionError as e_conn: 
        print(f"❌ CRITICAL ERROR: Could not connect to Bloomberg. {e_conn}")
        print("    Please ensure the Bloomberg Terminal is running, you are logged in, and the API is correctly configured (e.g., using 'bbcomm').")
    except ConnectionAbortedError as e_aborted: 
        print(f"❌ CRITICAL ERROR: The connection to Bloomberg was aborted during data fetching. {e_aborted}")
        print("    Some data might be missing. Please check the output file and consider re-running.")
    except blpapi.exception.BlpException as e_blp: 
        print(f"❌ CRITICAL BLPAPI ERROR: A Bloomberg specific error occurred: {e_blp}")
        print("    This could be due to various reasons like invalid security, field, or service issues.")
    except Exception as e_main:
        print(f"❌ AN UNEXPECTED CRITICAL ERROR occurred: {e_main}")
        import traceback
        print("\n--- Technical Error Details ---")
        traceback.print_exc()
        print("--- End Technical Error Details ---\n")
    finally:
        print("\n👋 Script execution finished. Have a great day!")
        print("\n    If you doubt any values from the report please double check with the terminal values!")
